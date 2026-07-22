import json
import os
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


ROOT = Path(__file__).resolve().parents[3]
PYTHON = ROOT / "python" / "venv" / "bin" / "python"
CONVERTER = ROOT / "python" / "pdf-editor" / "convert_to_excel.py"
LAYOUT_FIXTURE = ROOT / "tests" / "pdfjs" / "test_pdf_1.pdf"
DRYLAB_FIXTURE = ROOT / "tests" / "OverlayEditor" / "drylab.pdf"


class ExcelConversionTest(unittest.TestCase):
    def convert(self, source, output, *options):
        result = subprocess.run(
            [str(PYTHON), str(CONVERTER), str(source), str(output), *options, "--json"],
            cwd=ROOT,
            env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1"},
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        payloads = []
        for line in result.stdout.splitlines():
            try:
                payloads.append(json.loads(line))
            except json.JSONDecodeError:
                pass
        self.assertTrue(payloads, result.stdout)
        self.assertTrue(payloads[-1]["success"])
        self.assertTrue(output.is_file())
        return payloads[-1]

    def build_merged_table_pdf(self, path):
        document = SimpleDocTemplate(str(path), pagesize=letter)
        table = Table(
            [
                ["Quarterly Results", "", ""],
                ["Region", "Revenue", "Growth"],
                ["North", "$1,250.50", "12%"],
            ],
            colWidths=[120, 90, 90],
        )
        table.setStyle(TableStyle([
            ("SPAN", (0, 0), (2, 0)),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, 1), colors.lightgrey),
        ]))
        document.build([table])

    def test_tables_only_preserves_page_sheets_and_non_table_pages(self):
        with tempfile.TemporaryDirectory(prefix="toolbase_excel_test_") as directory:
            output = Path(directory) / "tables.xlsx"
            payload = self.convert(
                LAYOUT_FIXTURE,
                output,
                "--mode", "tables", "--merge-cells", "--sheet-per-page",
            )

            self.assertEqual(payload["tables_found"], 2)
            self.assertEqual(payload["sheets"], 3)
            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, ["Page 1", "Page 2", "Page 3"])
            expected = [[1, 2, 3, 4], [5, 6, 7, 8], [9, 10, 11, 12], [13, 14, 15, 16]]
            for title in ("Page 1", "Page 2"):
                actual = [[workbook[title].cell(row, column).value for column in range(1, 5)] for row in range(1, 5)]
                self.assertEqual(actual, expected)
            self.assertEqual(workbook["Page 3"]["A1"].value, "No tables detected on this page.")

    def test_zero_table_newsletter_preserves_reference_layout_structure(self):
        with tempfile.TemporaryDirectory(prefix="toolbase_excel_drylab_test_") as directory:
            output = Path(directory) / "drylab.xlsx"
            payload = self.convert(
                DRYLAB_FIXTURE,
                output,
                "--merge-cells", "--sheet-per-page",
            )

            self.assertEqual(payload["mode"], "all")
            self.assertEqual(payload["tables_found"], 0)
            self.assertFalse(payload["fallback_used"])
            self.assertEqual(payload["effective_mode"], "layout")
            self.assertEqual(payload["layout_engine"], "newsletter")
            self.assertEqual(payload["images_embedded"], 4)
            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, ["Table 1", "Table 2"])
            self.assertEqual(workbook["Table 1"].max_column, 3)
            self.assertEqual(workbook["Table 1"].max_row, 4)
            self.assertEqual(workbook["Table 2"].max_column, 2)
            self.assertEqual(workbook["Table 2"].max_row, 5)
            self.assertEqual(len(workbook["Table 1"]._images), 1)
            self.assertEqual(len(workbook["Table 2"]._images), 3)
            self.assertAlmostEqual(workbook["Table 1"].column_dimensions["A"].width, 15.555556, places=5)
            self.assertAlmostEqual(workbook["Table 1"].column_dimensions["B"].width, 41.333333, places=5)
            self.assertAlmostEqual(workbook["Table 1"].column_dimensions["C"].width, 60.222222, places=5)
            self.assertEqual(
                [workbook["Table 1"].row_dimensions[index].height for index in range(1, 5)],
                [19.5, 207.75, 100.5, 250.0],
            )
            self.assertEqual(
                [workbook["Table 2"].row_dimensions[index].height for index in range(1, 6)],
                [409.0, 307.0, 325.75, 36.0, 322.0],
            )
            self.assertEqual(
                {str(cell_range) for cell_range in workbook["Table 1"].merged_cells.ranges},
                {"A1:C1", "A2:B2", "B3:C3"},
            )
            self.assertEqual(
                {str(cell_range) for cell_range in workbook["Table 2"].merged_cells.ranges},
                {"A1:A2", "B1:B2", "A4:B4"},
            )
            self.assertEqual(workbook["Table 1"]["A1"].font.name, "Verdana")
            self.assertEqual(workbook["Table 1"]["A1"].font.size, 13)
            self.assertIsNone(workbook["Table 1"]["A1"].font.color)
            self.assertEqual(workbook["Table 1"]["A2"].font.name, "Times New Roman")
            self.assertEqual(workbook["Table 1"]["A2"].font.size, 10)
            self.assertEqual(workbook["Table 1"]["A2"].font.color.rgb, "FF000000")
            self.assertEqual(workbook["Table 1"]["B3"].font.name, "Tahoma")
            self.assertEqual(workbook["Table 1"]["B3"].font.size, 11.5)
            self.assertEqual(workbook["Table 2"]["A1"].font.name, "Times New Roman")
            self.assertEqual(workbook["Table 2"]["A1"].font.color.rgb, "FF000000")
            self.assertGreater(output.stat().st_size, 1_000_000)
            page_one_text = "\n".join(
                str(cell.value)
                for row in workbook["Table 1"].iter_rows()
                for cell in row
                if cell.value is not None
            )
            self.assertIn("Welcome to our first newsletter", page_one_text)
            self.assertIn("investment round that closed in January", page_one_text)
            self.assertNotIn("No tables detected", page_one_text)

    def test_source_merged_cells_and_numeric_formats_are_preserved(self):
        with tempfile.TemporaryDirectory(prefix="toolbase_excel_merge_test_") as directory:
            source = Path(directory) / "merged.pdf"
            output = Path(directory) / "merged.xlsx"
            self.build_merged_table_pdf(source)
            payload = self.convert(
                source, output, "--mode", "tables", "--merge-cells", "--sheet-per-page"
            )

            self.assertEqual(payload["merged_ranges"], 1)
            worksheet = load_workbook(output).active
            self.assertEqual({str(cell_range) for cell_range in worksheet.merged_cells.ranges}, {"A1:C1"})
            self.assertEqual(worksheet["B3"].value, 1250.5)
            self.assertEqual(worksheet["C3"].value, 0.12)
            self.assertEqual(worksheet["C3"].number_format, "0%")

            unmerged_output = Path(directory) / "unmerged.xlsx"
            self.convert(
                source, unmerged_output, "--mode", "tables", "--no-merge-cells", "--sheet-per-page"
            )
            self.assertEqual(list(load_workbook(unmerged_output).active.merged_cells.ranges), [])

    def test_all_content_mode_exports_positioned_text(self):
        with tempfile.TemporaryDirectory(prefix="toolbase_excel_all_test_") as directory:
            output = Path(directory) / "all.xlsx"
            payload = self.convert(
                LAYOUT_FIXTURE,
                output,
                "--mode", "all", "--no-merge-cells", "--single-sheet",
            )
            worksheet = load_workbook(output).active
            self.assertEqual(payload["images_embedded"], 1)
            self.assertEqual(len(worksheet._images), 1)
            self.assertTrue(
                any(
                    (dimension.height or 0) > 18
                    for dimension in worksheet.row_dimensions.values()
                )
            )
            lorem_cell = next(
                cell
                for row in worksheet.iter_rows()
                for cell in row
                if "Lorem ipsum dolor sit amet" in str(cell.value or "")
            )
            self.assertEqual(lorem_cell.font.name, "Open Sans")
            self.assertAlmostEqual(lorem_cell.font.size, 10.56, places=2)
            self.assertEqual(lorem_cell.font.color.rgb, "FF000000")
            text = "\n".join(
                str(cell.value)
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            self.assertIn("Lorem ipsum dolor sit amet", text)
            self.assertIn("Page 3", text)


if __name__ == "__main__":
    unittest.main()
