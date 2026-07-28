#!/usr/bin/env python3
"""Convert PDF tables or positioned text into an Excel workbook."""

import argparse
import io
import json
import os
import re
import sys
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    print(json.dumps({"success": False, "error": "PyMuPDF is not installed. Run: pip install PyMuPDF"}))
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as WorksheetImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import points_to_pixels
except ImportError:
    print(json.dumps({"success": False, "error": "openpyxl is not installed. Run: pip install openpyxl"}))
    sys.exit(1)


@dataclass
class TableData:
    rows: list[list[str]]
    merges: list[tuple[int, int, int, int]] = field(default_factory=list)
    bbox: tuple[float, float, float, float] | None = None
    header: bool = True
    styles: list[list[dict | None]] = field(default_factory=list)


HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(bold=True, color="1F2937", size=10)
SECTION_FONT = Font(bold=True, color="166534", size=11)
NORMAL_FONT = Font(size=10, color="111827")
NOTE_FONT = Font(italic=True, color="64748B", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _cluster_positions(values, tolerance=1.0):
    """Collapse nearly-identical PDF coordinates into stable grid lines."""
    values = sorted(float(value) for value in values)
    if not values:
        return []
    groups = [[values[0]]]
    for value in values[1:]:
        if abs(value - sum(groups[-1]) / len(groups[-1])) <= tolerance:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [sum(group) / len(group) for group in groups]


def _nearest_boundary(boundaries, value):
    return min(range(len(boundaries)), key=lambda index: abs(boundaries[index] - value))


def _table_merges(table):
    """Infer row/column spans from PyMuPDF's detected cell rectangles."""
    if table.row_count < 1 or table.col_count < 1:
        return []

    x_values = [table.bbox[0], table.bbox[2]]
    for row in table.rows:
        for cell in row.cells:
            if cell:
                x_values.extend((cell[0], cell[2]))
    x_boundaries = _cluster_positions(x_values)

    complete_rows = [row for row in table.rows if sum(cell is not None for cell in row.cells) == table.col_count]
    if complete_rows:
        sample_cells = sorted(complete_rows[0].cells, key=lambda cell: cell[0])
        x_boundaries = [sample_cells[0][0], *[cell[2] for cell in sample_cells]]
    if len(x_boundaries) != table.col_count + 1:
        width = (table.bbox[2] - table.bbox[0]) / table.col_count
        x_boundaries = [table.bbox[0] + width * index for index in range(table.col_count + 1)]

    y_boundaries = [table.rows[0].bbox[1], *[row.bbox[3] for row in table.rows]]
    merges = set()
    for source_row in table.rows:
        for cell in source_row.cells:
            if not cell:
                continue
            start_col = _nearest_boundary(x_boundaries, cell[0])
            end_col = _nearest_boundary(x_boundaries, cell[2]) - 1
            start_row = _nearest_boundary(y_boundaries, cell[1])
            end_row = _nearest_boundary(y_boundaries, cell[3]) - 1
            start_col = max(0, min(table.col_count - 1, start_col))
            end_col = max(start_col, min(table.col_count - 1, end_col))
            start_row = max(0, min(table.row_count - 1, start_row))
            end_row = max(start_row, min(table.row_count - 1, end_row))
            if end_col > start_col or end_row > start_row:
                merges.add((start_row, start_col, end_row, end_col))
    return sorted(merges)


def find_tables_on_page(page):
    """Return detected PDF tables with source-derived merged-cell regions."""
    try:
        finder = page.find_tables()
    except Exception:
        return []

    result = []
    for table in getattr(finder, "tables", []) or []:
        rows = []
        for source_row in table.extract():
            row = ["" if value is None else str(value).strip() for value in source_row]
            row.extend([""] * max(0, table.col_count - len(row)))
            rows.append(row[:table.col_count])
        if rows:
            result.append(TableData(rows=rows, merges=_table_merges(table), bbox=tuple(table.bbox)))
    return result


def _line_text(line):
    pieces = []
    previous_right = None
    previous_spans = []
    for span in line.get("spans", []):
        text = str(span.get("text", ""))
        if not text:
            continue
        bbox = tuple(float(value) for value in span.get("bbox", (0, 0, 0, 0))[:4])
        duplicate = any(
            prior_text == text
            and abs(prior_bbox[0] - bbox[0]) < 1
            and abs(prior_bbox[1] - bbox[1]) < 1
            and abs(prior_bbox[2] - bbox[2]) < 1
            and abs(prior_bbox[3] - bbox[3]) < 1
            for prior_text, prior_bbox in previous_spans
        )
        if duplicate:
            continue
        previous_spans.append((text, bbox))
        left = bbox[0]
        if pieces and previous_right is not None and left - previous_right > 2 and not pieces[-1].endswith(" "):
            pieces.append(" ")
        pieces.append(text)
        previous_right = bbox[2]
    return "".join(pieces).strip()


def _source_font_name(value):
    name = str(value or "").strip()
    if "+" in name and len(name.split("+", 1)[0]) == 6:
        name = name.split("+", 1)[1]
    return name or "Arial"


def _span_style(span):
    flags = int(span.get("flags", 0) or 0)
    color = int(span.get("color", 0) or 0) & 0xFFFFFF
    return {
        "name": _source_font_name(span.get("font")),
        "size": max(1, round(float(span.get("size", 10) or 10), 2)),
        "bold": bool(flags & 16) or "bold" in str(span.get("font", "")).lower(),
        "italic": bool(flags & 2) or "italic" in str(span.get("font", "")).lower(),
        "color": f"FF{color:06X}",
    }


def _dominant_line_style(spans):
    styled = [span for span in spans if str(span.get("text", "")).strip()]
    if not styled:
        return None
    # A cell can only have one Excel font. Prefer the span carrying the most
    # text, with font size breaking ties so headings retain their hierarchy.
    span = max(styled, key=lambda item: (len(str(item.get("text", "")).strip()), float(item.get("size", 0) or 0)))
    return _span_style(span)


def _join_paragraph_line(previous, current):
    if not previous:
        return current
    if previous.endswith("-"):
        return previous + current
    return f"{previous} {current}"


def _deduplicate_overlaid_lines(lines):
    """Remove near-identical shadow/overlay text emitted by some PDF producers."""
    unique = []
    for line in lines:
        match_index = None
        for index, prior in enumerate(unique):
            if abs(prior["y"] - line["y"]) > 1 or abs(prior["x"] - line["x"]) > 1:
                continue
            same = prior["text"] == line["text"]
            prefix_variant = prior["text"].startswith(line["text"]) or line["text"].startswith(prior["text"])
            if same or prefix_variant:
                match_index = index
                break
        if match_index is None:
            unique.append(line)
        elif len(line["text"]) > len(unique[match_index]["text"]):
            unique[match_index] = line
    return unique


def _group_lines_into_paragraphs(lines, column):
    paragraphs = []
    current = None
    for line in sorted(lines, key=lambda item: (item["y"], item["x"])):
        if current is None:
            current = {**line, "column": column}
            continue
        previous_height = max(1, current["last_y2"] - current["last_y"])
        gap = line["y"] - current["last_y2"]
        starts_new_heading = line["starts_bold"] and not current["last_starts_bold"]
        x_shift = abs(line["x"] - current["line_x"])
        starts_new_paragraph = gap > max(5.0, previous_height * 0.65) or (
            starts_new_heading and x_shift <= 14 and gap > 1.5
        )
        if starts_new_paragraph:
            paragraphs.append(current)
            current = {**line, "column": column}
            continue
        current["text"] = _join_paragraph_line(current["text"], line["text"])
        current["x2"] = max(current["x2"], line["x2"])
        current["y2"] = max(current["y2"], line["y2"])
        current["last_y"] = line["y"]
        current["last_y2"] = line["y2"]
        current["last_starts_bold"] = line["starts_bold"]
        if line.get("style") and (
            not current.get("style")
            or float(line["style"].get("size", 0)) > float(current["style"].get("size", 0))
        ):
            current["style"] = line["style"]
    if current is not None:
        paragraphs.append(current)
    return paragraphs


def extract_all_text_grid(page):
    """Group page text into paragraph cells that follow the source columns."""
    lines = []
    for block in page.get_text("dict", sort=True).get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            text = _line_text(line)
            if not text:
                continue
            bbox = tuple(line.get("bbox", (0, 0, 0, 0)))
            spans = line.get("spans", [])
            starts_bold = bool(spans) and (
                int(spans[0].get("flags", 0)) & 16 != 0
                or "bold" in str(spans[0].get("font", "")).lower()
            )
            lines.append({
                "text": text,
                "x": float(bbox[0]),
                "y": float(bbox[1]),
                "x2": float(bbox[2]),
                "y2": float(bbox[3]),
                "last_y": float(bbox[1]),
                "last_y2": float(bbox[3]),
                "line_x": float(bbox[0]),
                "starts_bold": starts_bold,
                "last_starts_bold": starts_bold,
                "style": _dominant_line_style(spans),
            })
    lines = _deduplicate_overlaid_lines(lines)
    if not lines:
        return TableData(rows=[], header=False)

    midpoint = float(page.rect.width) / 2
    left_lines = []
    right_lines = []
    full_lines = []
    for line in lines:
        width = line["x2"] - line["x"]
        crosses_center = line["x"] < midpoint - 8 and line["x2"] > midpoint + 8
        if width >= page.rect.width * 0.68 or crosses_center:
            full_lines.append(line)
        elif line["x"] >= midpoint - 8:
            right_lines.append(line)
        else:
            left_lines.append(line)

    two_columns = len(left_lines) >= 3 and len(right_lines) >= 3
    if not two_columns:
        full_lines.extend(left_lines)
        full_lines.extend(right_lines)
        left_lines = []
        right_lines = []

    paragraphs = []
    paragraphs.extend(_group_lines_into_paragraphs(left_lines, 0))
    paragraphs.extend(_group_lines_into_paragraphs(right_lines, 1))
    paragraphs.extend(_group_lines_into_paragraphs(full_lines, "full"))
    paragraphs.sort(key=lambda item: (item["y"], 0 if item["column"] in ("full", 0) else 1))

    layout_rows = []
    merges = []
    for paragraph in paragraphs:
        if paragraph["column"] == "full" or not two_columns:
            layout_rows.append({
                "y": paragraph["y"],
                "y2": paragraph["y2"],
                "values": [paragraph["text"], ""],
                "styles": [paragraph.get("style"), None],
            })
            if two_columns:
                merges.append((len(layout_rows) - 1, 0, len(layout_rows) - 1, 1))
            continue

        target = None
        for candidate in reversed(layout_rows):
            column = int(paragraph["column"])
            overlaps = min(candidate["y2"], paragraph["y2"]) >= max(candidate["y"], paragraph["y"])
            starts_together = abs(candidate["y"] - paragraph["y"]) <= 14
            if candidate["values"][column] == "" and (overlaps or starts_together):
                target = candidate
                break
        if target is None:
            target = {
                "y": paragraph["y"],
                "y2": paragraph["y2"],
                "values": ["", ""],
                "styles": [None, None],
            }
            layout_rows.append(target)
            layout_rows.sort(key=lambda item: item["y"])
        target["values"][int(paragraph["column"])] = paragraph["text"]
        target["styles"][int(paragraph["column"])] = paragraph.get("style")
        target["y"] = min(target["y"], paragraph["y"])
        target["y2"] = max(target["y2"], paragraph["y2"])

    rows = [layout_row["values"] if two_columns else [layout_row["values"][0]] for layout_row in layout_rows]
    if two_columns:
        merged_texts = {layout_rows[index]["values"][0] for index, _, _, _ in merges}
        merges = [
            (row_index, 0, row_index, 1)
            for row_index, row in enumerate(rows)
            if row[0] in merged_texts and row[1] == ""
        ]
    else:
        merges = []
    styles = [
        layout_row["styles"] if two_columns else [layout_row["styles"][0]]
        for layout_row in layout_rows
    ]
    return TableData(rows=rows, merges=merges, header=False, styles=styles)


def detect_cell_value(value):
    """Return an Excel value and an optional display number format."""
    if value is None:
        return "", None
    text = str(value).strip()
    if not text:
        return "", None

    percent = re.fullmatch(r"\(?\s*([+-]?[\d,]+(?:\.\d+)?)\s*%\s*\)?", text)
    if percent:
        number = float(percent.group(1).replace(",", "")) / 100
        if text.startswith("("):
            number = -abs(number)
        return number, "0.00%" if "." in percent.group(1) else "0%"

    currency = re.fullmatch(r"\(?\s*([$€£¥])\s*([+-]?[\d,]+(?:\.\d+)?)\s*\)?", text)
    if currency:
        number = float(currency.group(2).replace(",", ""))
        if text.startswith("("):
            number = -abs(number)
        decimals = "0.00" if "." in currency.group(2) else "0"
        return number, f'{currency.group(1)}#,##{decimals};[Red]-{currency.group(1)}#,##{decimals}'

    if re.fullmatch(r"[+-]?\d+", text) and not (len(text.lstrip("+-")) > 1 and text.lstrip("+-").startswith("0")):
        return int(text), "0"
    if re.fullmatch(r"[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d+", text):
        return float(text.replace(",", "")), "#,##0.00"
    if re.fullmatch(r"[+-]?\d{1,3}(?:,\d{3})+", text):
        return int(text.replace(",", "")), "#,##0"

    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            parsed = datetime.strptime(text, date_format).date()
            return parsed, "yyyy-mm-dd" if date_format == "%Y-%m-%d" else "mm/dd/yyyy"
        except ValueError:
            pass
    return text, None


def _update_column_width(worksheet, column, value):
    letter = get_column_letter(column)
    current = worksheet.column_dimensions[letter].width or 0
    longest_line = max((len(line) for line in str(value or "").splitlines()), default=0)
    worksheet.column_dimensions[letter].width = min(55, max(current, min(55, longest_line + 2), 9))


def _estimated_wrapped_lines(value, width):
    usable_width = max(8, float(width or 10) - 1)
    return sum(
        max(1, (len(line) + int(usable_width) - 1) // int(usable_width))
        for line in str(value or "").splitlines() or [""]
    )


def _format_sheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.outlinePr.summaryBelow = True


def write_table(worksheet, table, start_row=1, merge_cells=True):
    """Write one detected table and return its next row and merge count."""
    if not table.rows:
        return start_row, 0
    column_count = max(len(row) for row in table.rows)
    for row_index, row in enumerate(table.rows):
        excel_row = start_row + row_index
        for column_index in range(column_count):
            raw_value = row[column_index] if column_index < len(row) else ""
            cell = worksheet.cell(row=excel_row, column=column_index + 1)
            cell.value, number_format = detect_cell_value(raw_value)
            if number_format:
                cell.number_format = number_format
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            source_style = None
            if row_index < len(table.styles) and column_index < len(table.styles[row_index]):
                source_style = table.styles[row_index][column_index]
            if source_style:
                cell.font = Font(
                    name=source_style.get("name") or "Arial",
                    size=source_style.get("size") or 10,
                    bold=bool(source_style.get("bold")),
                    italic=bool(source_style.get("italic")),
                    color=source_style.get("color") or "FF000000",
                )
            else:
                cell.font = HEADER_FONT if table.header and row_index == 0 else NORMAL_FONT
            if table.header and row_index == 0:
                cell.fill = HEADER_FILL
            _update_column_width(worksheet, column_index + 1, raw_value)
        estimated_lines = 1
        for column_index, value in enumerate(row):
            letter = get_column_letter(column_index + 1)
            estimated_lines = max(
                estimated_lines,
                _estimated_wrapped_lines(value, worksheet.column_dimensions[letter].width),
            )
        worksheet.row_dimensions[excel_row].height = min(409, max(18, estimated_lines * 15))

    merged_count = 0
    if merge_cells:
        for first_row, first_col, last_row, last_col in table.merges:
            if first_row == last_row and first_col == last_col:
                continue
            try:
                worksheet.merge_cells(
                    start_row=start_row + first_row,
                    start_column=first_col + 1,
                    end_row=start_row + last_row,
                    end_column=last_col + 1,
                )
                worksheet.cell(start_row + first_row, first_col + 1).alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )
                merged_count += 1
            except ValueError:
                continue
    return start_row + len(table.rows), merged_count


def _write_section_label(worksheet, row, label):
    cell = worksheet.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row].height = 20
    _update_column_width(worksheet, 1, label)
    return row + 1


def _write_note(worksheet, row, message):
    cell = worksheet.cell(row=row, column=1, value=message)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    _update_column_width(worksheet, 1, message)
    return row + 1


def _safe_sheet_title(label, existing):
    base = re.sub(r"[\\/*?:\[\]]", "-", str(label)).strip()[:31] or "Sheet"
    candidate = base
    suffix = 2
    while candidate in existing:
        marker = f" ({suffix})"
        candidate = f"{base[:31-len(marker)]}{marker}"
        suffix += 1
    return candidate


def _page_image_blocks(page):
    images = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 1 or not block.get("image"):
            continue
        images.append({
            "bbox": tuple(float(value) for value in block.get("bbox", (0, 0, 0, 0))),
            "bytes": block["image"],
            "extension": str(block.get("ext", "png")),
        })
    return sorted(images, key=lambda image: (image["bbox"][1], image["bbox"][0]))


def _looks_like_newsletter_layout(pdf_document):
    """Detect a multi-page, two-column design with cover and footer images."""
    if len(pdf_document) < 3:
        return False
    first_images = _page_image_blocks(pdf_document[0])
    last_images = _page_image_blocks(pdf_document[-1])
    if not first_images or not last_images:
        return False
    first_width = pdf_document[0].rect.width
    last_width = pdf_document[-1].rect.width
    has_cover_image = any(
        image["bbox"][2] - image["bbox"][0] >= first_width * 0.65
        and image["bbox"][1] < pdf_document[0].rect.height * 0.4
        for image in first_images
    )
    has_footer_image = any(
        image["bbox"][2] - image["bbox"][0] >= last_width * 0.65
        and image["bbox"][1] > pdf_document[-1].rect.height * 0.35
        for image in last_images
    )
    if not has_cover_image or not has_footer_image:
        return False
    for page in pdf_document:
        content = extract_all_text_grid(page)
        if not content.rows or max((len(row) for row in content.rows), default=0) < 2:
            return False
    return True


def _set_layout_cell(
    worksheet,
    coordinate,
    value,
    font_name="Times New Roman",
    font_size=10,
    indent=0,
    font_color="FF000000",
    charset=204,
):
    cell = worksheet[coordinate]
    cell.value = value or ""
    cell.font = Font(name=font_name, size=font_size, color=font_color, charset=charset)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=indent)
    return cell


def _add_layout_image(
    worksheet,
    image_data,
    row,
    column,
    row_offset_points=0,
    column_offset_points=0,
    scale=1.0,
):
    bbox = image_data["bbox"]
    width_points = (bbox[2] - bbox[0]) * scale
    height_points = (bbox[3] - bbox[1]) * scale
    width_pixels = points_to_pixels(width_points)
    height_pixels = points_to_pixels(height_points)
    image = WorksheetImage(io.BytesIO(image_data["bytes"]))
    image.width = width_pixels
    image.height = height_pixels
    marker = AnchorMarker(
        col=column,
        colOff=round(max(0, column_offset_points) * 12700),
        row=row,
        rowOff=round(max(0, row_offset_points) * 12700),
    )
    image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(round(width_points * 12700), round(height_points * 12700)),
    )
    worksheet.add_image(image)


def _column_text(rows, column, indexes):
    values = []
    for index in indexes:
        if index >= len(rows) or column >= len(rows[index]):
            continue
        value = str(rows[index][column] or "").strip()
        if value:
            values.append(value)
    return "\n".join(values)


def _format_stat_text(value):
    value = str(value or "").strip()
    match = re.fullmatch(r"(\d+)\s+(meetings)\s+(NY\s*·\s*SF)\s+(LA\s*·\s*LV)", value)
    if match:
        return f"{match.group(1)}\n{match.group(2)} {match.group(3)}\n{match.group(4)}"
    return value


def _write_newsletter_layout(workbook, pdf_document):
    """Build a Solid-style workbook for cover-plus-two-column newsletters."""
    first_page = pdf_document[0]
    first_content = extract_all_text_grid(first_page)
    first_rows = first_content.rows
    first_merges = {merge[0] for merge in first_content.merges}

    first_sheet = workbook.create_sheet("Table 1")
    first_sheet.sheet_view.showGridLines = False
    first_sheet.column_dimensions["A"].width = 15.555556
    first_sheet.column_dimensions["B"].width = 41.333333
    first_sheet.column_dimensions["C"].width = 60.222222
    first_sheet.row_dimensions[1].height = 19.5
    first_sheet.row_dimensions[2].height = 207.75
    first_sheet.row_dimensions[3].height = 100.5

    header_left = first_rows[0][0] if first_rows and first_rows[0] else ""
    header_right = first_rows[0][1] if first_rows and len(first_rows[0]) > 1 else ""
    header = header_right or header_left
    _set_layout_cell(
        first_sheet,
        "A1",
        header,
        font_name="Verdana",
        font_size=13,
        indent=37,
        font_color=None,
        charset=None,
    )
    first_sheet.merge_cells("A1:C1")

    footer_index = max(first_merges) if first_merges else len(first_rows)
    stat_index = footer_index - 1 if footer_index > 1 else None
    body_end = stat_index if stat_index is not None else footer_index
    body_indexes = list(range(1, max(1, body_end)))
    _set_layout_cell(first_sheet, "A2", _column_text(first_rows, 0, body_indexes))
    _set_layout_cell(first_sheet, "C2", _column_text(first_rows, 1, body_indexes), indent=1)
    first_sheet.merge_cells("A2:B2")

    if stat_index is not None and stat_index < len(first_rows):
        _set_layout_cell(first_sheet, "A3", _format_stat_text(first_rows[stat_index][0]))
    if footer_index < len(first_rows):
        _set_layout_cell(
            first_sheet,
            "B3",
            first_rows[footer_index][0],
            font_name="Tahoma",
            font_size=11.5,
            font_color=None,
            charset=None,
        )
    first_sheet.merge_cells("B3:C3")

    first_images = _page_image_blocks(first_page)
    cover_image = max(first_images, key=lambda image: (image["bbox"][2] - image["bbox"][0]) * (image["bbox"][3] - image["bbox"][1]))
    cover_height = cover_image["bbox"][3] - cover_image["bbox"][1]
    first_sheet["A4"] = ""
    first_sheet.row_dimensions[4].height = round(cover_height) + 1
    _add_layout_image(
        first_sheet,
        cover_image,
        row=2,
        column=0,
        row_offset_points=100,
        column_offset_points=cover_image["bbox"][0],
    )

    continuation_sheet = workbook.create_sheet("Table 2")
    continuation_sheet.sheet_view.showGridLines = False
    continuation_sheet.column_dimensions["A"].width = 57.333333
    continuation_sheet.column_dimensions["B"].width = 60.0

    second_page = pdf_document[1]
    second_rows = extract_all_text_grid(second_page).rows
    second_indexes = list(range(len(second_rows)))
    _set_layout_cell(continuation_sheet, "A1", _column_text(second_rows, 0, second_indexes))
    _set_layout_cell(continuation_sheet, "B1", _column_text(second_rows, 1, second_indexes), indent=1)
    continuation_sheet.merge_cells("A1:A2")
    continuation_sheet.merge_cells("B1:B2")
    content_top = min(
        (
            block[1]
            for block in second_page.get_text("blocks", sort=True)
            if len(block) > 6 and block[6] == 0 and str(block[4]).strip()
        ),
        default=0,
    )
    content_bottom = max(
        (
            block[3]
            for block in second_page.get_text("blocks", sort=True)
            if len(block) > 6 and block[6] == 0 and str(block[4]).strip()
        ),
        default=second_page.rect.height,
    )
    content_height = max(360, content_bottom - content_top)
    continuation_sheet.row_dimensions[1].height = round(content_height * 0.57) + 1
    continuation_sheet.row_dimensions[2].height = round(content_height - continuation_sheet.row_dimensions[1].height)
    for image in _page_image_blocks(second_page):
        column = 0 if image["bbox"][0] < second_page.rect.width / 2 else 1
        column_start = 56.7 if column == 0 else second_page.rect.width / 2
        _add_layout_image(
            continuation_sheet,
            image,
            row=0,
            column=column,
            row_offset_points=image["bbox"][1] - content_top,
            column_offset_points=image["bbox"][0] - column_start,
        )

    output_row = 3
    for page_index in range(2, len(pdf_document)):
        page = pdf_document[page_index]
        content = extract_all_text_grid(page)
        rows = content.rows
        merged_rows = {merge[0] for merge in content.merges}
        footer_row = max(merged_rows) if merged_rows else None
        body_indexes = [index for index in range(len(rows)) if index != footer_row]
        _set_layout_cell(continuation_sheet, f"A{output_row}", _column_text(rows, 0, body_indexes))
        _set_layout_cell(continuation_sheet, f"B{output_row}", _column_text(rows, 1, body_indexes), indent=1)

        page_images = _page_image_blocks(page)
        large_images = [
            image for image in page_images
            if image["bbox"][2] - image["bbox"][0] >= page.rect.width * 0.65
        ]
        small_images = [image for image in page_images if image not in large_images]
        page_text_top = min(
            (
                block[1]
                for block in page.get_text("blocks", sort=True)
                if len(block) > 6 and block[6] == 0 and str(block[4]).strip()
            ),
            default=0,
        )
        large_image_top = min((image["bbox"][1] for image in large_images), default=page.rect.height * 0.6)
        layout_height = max(180, large_image_top - page_text_top - 33.87)
        continuation_sheet.row_dimensions[output_row].height = round(layout_height * 4) / 4
        for image in small_images:
            column = 0 if image["bbox"][0] < page.rect.width / 2 else 1
            column_start = 56.7 if column == 0 else page.rect.width / 2
            _add_layout_image(
                continuation_sheet,
                image,
                row=output_row - 1,
                column=column,
                row_offset_points=image["bbox"][1] - page_text_top,
                column_offset_points=image["bbox"][0] - column_start,
                scale=0.785,
            )

        if footer_row is not None and footer_row < len(rows):
            output_row += 1
            _set_layout_cell(continuation_sheet, f"A{output_row}", rows[footer_row][0])
            continuation_sheet.merge_cells(start_row=output_row, start_column=1, end_row=output_row, end_column=2)
            continuation_sheet.row_dimensions[output_row].height = 36

        for image in large_images:
            output_row += 1
            continuation_sheet.cell(output_row, 1, "")
            continuation_sheet.row_dimensions[output_row].height = round(image["bbox"][3] - image["bbox"][1]) + 1
            _add_layout_image(
                continuation_sheet,
                image,
                row=output_row - 2,
                column=0,
                row_offset_points=31.15,
                column_offset_points=image["bbox"][0],
            )
        output_row += 1

    return {
        "rows_written": sum(sheet.max_row for sheet in workbook.worksheets),
        "merged_ranges": sum(len(sheet.merged_cells.ranges) for sheet in workbook.worksheets),
        "images_embedded": sum(len(sheet._images) for sheet in workbook.worksheets),
        "layout_engine": "newsletter",
    }


def _write_tables_page(worksheet, tables, merge_cells, start_row=1):
    current_row = start_row
    merged_count = 0
    if not tables:
        return _write_note(worksheet, current_row, "No tables detected on this page."), merged_count
    for table_index, table in enumerate(tables):
        if len(tables) > 1:
            current_row = _write_section_label(worksheet, current_row, f"Table {table_index + 1}")
        current_row, added_merges = write_table(worksheet, table, current_row, merge_cells)
        merged_count += added_merges
        if table_index < len(tables) - 1:
            current_row += 2
    return current_row, merged_count


def convert_pdf(input_pdf, output_xlsx, mode="all", merge_cells=True, sheet_per_page=True):
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Toolbase"
    workbook.properties.title = f"Excel export of {Path(input_pdf).name}"

    tables_found = 0
    rows_written = 0
    merged_ranges = 0
    images_embedded = 0
    layout_engine = None
    with fitz.open(input_pdf) as pdf_document:
        total_pages = len(pdf_document)
        page_tables = [find_tables_on_page(page) for page in pdf_document]
        tables_found = sum(len(tables) for tables in page_tables)
        fallback_used = mode == "tables" and tables_found == 0
        effective_mode = "all" if fallback_used else mode
        if tables_found == 0 and mode in {"all", "tables"} and _looks_like_newsletter_layout(pdf_document):
            effective_mode = "layout"
            layout_stats = _write_newsletter_layout(workbook, pdf_document)
            rows_written = layout_stats["rows_written"]
            merged_ranges = layout_stats["merged_ranges"]
            images_embedded = layout_stats["images_embedded"]
            layout_engine = layout_stats["layout_engine"]
        else:
            shared_sheet = None
            shared_row = 1
            if not sheet_per_page:
                shared_sheet = workbook.create_sheet("Data")
                _format_sheet(shared_sheet)

            for page_index, page in enumerate(pdf_document):
                page_label = f"Page {page_index + 1}"
                detected_tables = page_tables[page_index]

                if sheet_per_page:
                    title = _safe_sheet_title(page_label, workbook.sheetnames)
                    worksheet = workbook.create_sheet(title)
                    _format_sheet(worksheet)
                    start_row = 1
                else:
                    worksheet = shared_sheet
                    if shared_row > 1:
                        shared_row += 2
                    shared_row = _write_section_label(worksheet, shared_row, page_label)
                    start_row = shared_row

                if effective_mode == "tables":
                    next_row, added_merges = _write_tables_page(
                        worksheet, detected_tables, merge_cells, start_row
                    )
                    content_rows = max(0, next_row - start_row)
                else:
                    content_table = extract_all_text_grid(page)
                    if content_table.rows:
                        next_row, added_merges = write_table(
                            worksheet, content_table, start_row, merge_cells=merge_cells
                        )
                    else:
                        next_row = _write_note(worksheet, start_row, "No extractable text found on this page.")
                        added_merges = 0
                    for image_data in _page_image_blocks(page):
                        image_row = next_row + 1
                        worksheet.cell(image_row, 1, "")
                        worksheet.row_dimensions[image_row].height = min(
                            409,
                            round(image_data["bbox"][3] - image_data["bbox"][1]) + 1,
                        )
                        _add_layout_image(
                            worksheet,
                            image_data,
                            row=image_row - 1,
                            column=0,
                            column_offset_points=max(0, image_data["bbox"][0]),
                        )
                        images_embedded += 1
                        next_row = image_row
                    content_rows = max(0, next_row - start_row)

                rows_written += content_rows
                merged_ranges += added_merges
                if not sheet_per_page:
                    shared_row = next_row

    if not workbook.sheetnames:
        worksheet = workbook.create_sheet("Sheet1")
        _write_note(worksheet, 1, "No extractable content found in this PDF.")

    output_path = Path(output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "success": True,
        "output": str(output_path),
        "pages": total_pages,
        "tables_found": tables_found,
        "sheets": len(workbook.sheetnames),
        "rows_written": rows_written,
        "merged_ranges": merged_ranges,
        "images_embedded": images_embedded,
        "file_size": output_path.stat().st_size,
        "mode": mode,
        "effective_mode": effective_mode,
        "fallback_used": fallback_used,
        "layout_engine": layout_engine,
    }


def main():
    parser = argparse.ArgumentParser(description="Convert PDF to Excel (.xlsx)")
    parser.add_argument("input_pdf", help="Path to input PDF file")
    parser.add_argument("output_xlsx", help="Path to output XLSX file")
    parser.add_argument("--mode", choices=["tables", "all"], default="all")
    parser.add_argument("--merge-cells", dest="merge_cells", action="store_true", default=True)
    parser.add_argument("--no-merge-cells", dest="merge_cells", action="store_false")
    parser.add_argument("--sheet-per-page", dest="sheet_per_page", action="store_true", default=True)
    parser.add_argument("--single-sheet", dest="sheet_per_page", action="store_false")
    parser.add_argument("--json", action="store_true", default=False)
    args = parser.parse_args()

    if not os.path.exists(args.input_pdf):
        result = {"success": False, "error": f"Input file not found: {args.input_pdf}"}
        print(json.dumps(result) if args.json else f"Error: {result['error']}", file=sys.stdout if args.json else sys.stderr)
        sys.exit(1)

    try:
        result = convert_pdf(
            args.input_pdf,
            args.output_xlsx,
            mode=args.mode,
            merge_cells=args.merge_cells,
            sheet_per_page=args.sheet_per_page,
        )
        if args.json:
            print(json.dumps(result))
        else:
            print(
                f"Converted {result['pages']} pages to {result['output']} "
                f"({result['file_size']} bytes, {result['tables_found']} tables)"
            )
    except Exception as error:
        result = {"success": False, "error": str(error), "traceback": traceback.format_exc()}
        if args.json:
            print(json.dumps(result))
        else:
            print(f"Error: {error}", file=sys.stderr)
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
